# -*- coding:utf-8 -*-

# Mixin para exportar e importar de forma generica archivos csv, xls y xlsx

import csv
import xlrd
from openpyxl import load_workbook, reader
from utils import DictUnicodeWriter

class ExImporter(object):
    """
    Carga un archivo tipo xls, csv, xlsx, etc y lo intenta importar como una
    lista de valores para un modelo.

    Ejemplo:

    >>> from eximporter import eximporter
    >>> fname = '.some_excel_data.xls'
    >>> mi = ExImporter()
    >>> mi.load_file(fname)
    >>> 
    >>> for data in mi.data:
    >>>     register = mi.migrate_columns(data, dcolumns)

    """
    data = []

    def __init__(self, *args, **kwargs): 
        self.data = []

    def load_file(self, fname):
        """
        Instancia el archivo para iterarlo
        """
        self.data = self.get_data(fname)
        
    def get_data(self, fname):
        """
        Carga el formato correcto del archivo y devuelve un iterador con la
        lista de datos obtenidos
        """
        # intenta cargar como xls

        try:
            wb = load_workbook(filename=fname)

        except reader.excel.InvalidFileException:
            try:
                wb = xlrd.open_workbook(filename=fname)

            except UnicodeDecodeError:
                wb = xlrd.open_workbook(filename=fname, encoding_override="cp1252")

            except xlrd.XLRDError:
                csvfile = open(fname, 'Urb')
                try:
                    dialect = csv.Sniffer().sniff(csvfile.read(1024), delimiters=',')
                    csvfile.seek(0)
                    wb = csv.reader(csvfile, dialect)
                except:
                    csvfile.seek(0)
                    wb = csv.reader(csvfile)

                # procesa el CVS
                for row in wb:
                    yield row

            else:
                # procesa el xlrd
                sheet = wb.sheet_by_index(0)

                for rindex in range(sheet.nrows):
                    data = [value for value in sheet.row_values(rindex)]
                    yield data

        else:
            # procesa load_workbook
            sheet = wb.active
            
            for row in sheet.iter_rows():
                data = [cell.value for cell in row]
                yield data

    def get_columns(self):
        """
        Retorna las columnas cargadas del archivo
        """
        data = self.data.next()
        return dict((i, v) for i, v in enumerate(data))

    def migrate_columns(self, data, dcolumns):
        """
        Crea un map de las columnas con los datos de la lista y devuelve un 
        dict con el mapping.

        data = [[...], [...]]
        columns = { 0: 'column name', 1: 'column name...', ...}

        """
        return dict(
                [
                    (dcolumns[str(index)], value) \
                            for index, value in enumerate(data) \
                            if dcolumns.has_key(str(index))
                ]
            )

    def export(self, fname, dcolumns):
        """
        Exporta a csv el archivo cargado
        """
        fieldnames = dcolumns.values()
        with open(fname, 'wb') as csvfile:
            csvfile.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
            writer = DictUnicodeWriter(csvfile, fieldnames)
            writer.writeheader()

            for data in self.data:
                writer.writerow(self.migrate_columns(data, dcolumns))

            csvfile.close()
